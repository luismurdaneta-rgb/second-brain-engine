#!/usr/bin/env python3
"""auto_curate_folder.py — Convert a RAW/<folder>/ into Vaults/<folder>/.

Usage (from this sandbox or your Mac):
    python3 auto_curate_folder.py "<folder name under RAW/>"
    python3 auto_curate_folder.py "Recibos verdes"
    python3 auto_curate_folder.py --raw "/path/to/your/second-brain/RAW " "Recibos verdes"
    python3 auto_curate_folder.py --vaults "/path/to/your/second-brain/Vaults" "Bret and Rachel"

What it does:
  1. Walks RAW/<folder>/ recursively.
  2. Creates Vaults/<folder>/ if missing.
  3. For each file:
       - Always writes a _Sources/<safe-name>.md stub (raw_path, raw_path_uri, kind, size).
       - For PDFs: native text -> Notes/; if no text layer, OCR fallback (scanned
         permits, stamped drawings) unless --no-ocr.
       - For .docx: extracts text into Notes/<safe-name>.md.
       - For .txt/.md: copies into Notes/.
       - For images (png/jpg/...): OCR -> Notes/ if text is found, else source-only.
       - For .csv/.xlsx/.json: source stub + table/structured preview note.
       - For CAD/binary/media: source stub only.
  4. Concept links: rule-based candidate extraction (stopword-filtered, deduped),
     optionally refined by an LLM pass with --use-ai (needs `anthropic` +
     ANTHROPIC_API_KEY; off by default so a default run is fully local/offline).
  5. Cross-links: Notes/<x>.md gets a Sources block pointing to its _Sources stub.
  6. Builds _Index.md MOC + .obsidian/graph.json color groups.
  7. Writes _Ingestion-Report.md: per-file outcome (extracted / OCR / source-only /
     failed+reason) so nothing fails silently — the client trust layer.
  8. Cross-vault hint: scans other vaults for any note whose stem matches a new
     filename, prints a hint (no automatic edits to other vaults).

Key flags: --no-ocr (disable OCR), --use-ai (LLM concept refinement),
           --ai-model, --overwrite-notes, --vault-name.

Idempotent: safe to re-run; existing _Sources stubs are overwritten, Notes are
preserved unless --overwrite-notes is passed.
"""
from __future__ import annotations
import argparse
import json
import os
import re
import shutil
import sys
import textwrap
from collections import Counter, defaultdict
from pathlib import Path
from urllib.parse import quote

import lib_vault

# ---- Paths (edit if your Second Brain lives elsewhere) ----
DEFAULT_BASE = lib_vault.HOST_BASE
DEFAULT_RAW = DEFAULT_BASE / "RAW "  # NOTE: trailing space matches user's folder
DEFAULT_VAULTS = DEFAULT_BASE / "Vaults"

# ---- File-kind catalog ----
EXT_KINDS = {
    "pdf": ("pdf", "📄", "PDF document"),
    "docx": ("doc", "📝", "Word document"),
    "doc":  ("doc", "📝", "Word document"),
    "xlsx": ("spreadsheet", "📊", "Spreadsheet"),
    "xls":  ("spreadsheet", "📊", "Spreadsheet"),
    "csv":  ("data", "📊", "CSV data"),
    "tsv":  ("data", "📊", "TSV data"),
    "txt":  ("text", "📃", "Text"),
    "md":   ("markdown", "📝", "Markdown"),
    "py":   ("code", "🐍", "Python source"),
    "js":   ("code", "📜", "JavaScript source"),
    "ts":   ("code", "📜", "TypeScript source"),
    "ipynb":("notebook", "📓", "Jupyter notebook"),
    "json": ("data", "🗂", "JSON data"),
    "yaml": ("data", "🗂", "YAML data"),
    "yml":  ("data", "🗂", "YAML data"),
    "jpg":  ("image", "🖼", "Image"),
    "jpeg": ("image", "🖼", "Image"),
    "png":  ("image", "🖼", "Image"),
    "webp": ("image", "🖼", "Image"),
    "gif":  ("image", "🖼", "Image"),
    "svg":  ("image", "🖼", "Vector image"),
    "dwg":  ("cad", "📐", "AutoCAD drawing"),
    "rvt":  ("cad", "📐", "Revit project"),
    "rfa":  ("cad", "📐", "Revit family"),
    "ifc":  ("cad", "📐", "IFC BIM model"),
    "rws":  ("cad", "📐", "Revit workshare"),
    "dat":  ("binary", "💾", "Binary data"),
    "mp3":  ("audio", "🎵", "Audio"),
    "wav":  ("audio", "🎵", "Audio"),
    "mp4":  ("video", "🎬", "Video"),
    "mov":  ("video", "🎬", "Video"),
    "zip":  ("archive", "🗜", "Archive"),
}
DEFAULT_KIND = ("file", "📎", "File")

# Extensions we extract text from
TEXT_EXTRACT = {"pdf", "docx", "doc", "txt", "md", "ipynb", "csv", "tsv",
                "xlsx", "xls", "json", "yaml", "yml", "py", "js", "ts"}
SOURCE_ONLY = {"jpg","jpeg","png","webp","gif","svg","dwg","rvt","rfa","ifc","rws","dat","mp3","wav","mp4","mov","zip","opus","m4a"}
# Raster images we'll *attempt* OCR on when --ocr is enabled (default on).
# If OCR finds text, a Note is produced; otherwise the file stays source-only.
OCR_IMAGE = {"png", "jpg", "jpeg", "webp", "tiff", "tif", "bmp", "gif"}


# PDFs larger than this are stubbed as source-only (extraction would stall).
LARGE_PDF_BYTES = 8 * 1024 * 1024  # 8 MB


def safe(name: str, max_len: int = 120) -> str:
    s = re.sub(r'[/\\:*?"<>|]', "-", name)
    s = re.sub(r"\s+", " ", s).strip().strip(".")
    return (s or "Untitled")[:max_len]


def file_uri(macos_abs: str) -> str:
    p = macos_abs.replace("\\", "/")
    if not p.startswith("/"):
        p = "/" + p
    return "file://" + quote(p, safe="/:")


def kind_for(ext: str):
    return EXT_KINDS.get(ext.lower().lstrip("."), DEFAULT_KIND)


# Match ANY Cowork sandbox prefix, not just one defunct session name.
_SANDBOX_PREFIX_RE = re.compile(r"^/sessions/[^/]+/mnt/Second Brain")
HOST_VAULT_ROOT = str(lib_vault.HOST_BASE)


def to_macos_path(p: Path, raw_root_macos: Path | None = None) -> str:
    """Translate sandboxed path back to user's macOS path.

    Cowork mounts the vault under /sessions/<session>/mnt/Second Brain/. Each
    session name is ephemeral, so any literal sandbox name will break on the
    next run. We canonicalize with a regex instead of hardcoding session names.
    """
    return _SANDBOX_PREFIX_RE.sub(HOST_VAULT_ROOT, str(p))


def extract_pdf_text(path: Path) -> str | None:
    try:
        import pdfplumber
        text_parts = []
        with pdfplumber.open(path) as pdf:
            for i, page in enumerate(pdf.pages, start=1):
                t = page.extract_text() or ""
                if t.strip():
                    text_parts.append(f"### Page {i}\n\n{t.strip()}")
        return "\n\n".join(text_parts) if text_parts else None
    except Exception as e:
        try:
            import pypdf
            r = pypdf.PdfReader(str(path))
            parts = []
            for i, page in enumerate(r.pages, start=1):
                t = page.extract_text() or ""
                if t.strip():
                    parts.append(f"### Page {i}\n\n{t.strip()}")
            return "\n\n".join(parts) if parts else None
        except Exception:
            return None


def extract_docx_text(path: Path) -> str | None:
    try:
        import docx
        d = docx.Document(str(path))
        parts = []
        for para in d.paragraphs:
            if para.text.strip():
                style = para.style.name if para.style else ""
                if style.startswith("Heading 1"):   parts.append(f"# {para.text}")
                elif style.startswith("Heading 2"): parts.append(f"## {para.text}")
                elif style.startswith("Heading 3"): parts.append(f"### {para.text}")
                else:                                parts.append(para.text)
        return "\n\n".join(parts) if parts else None
    except Exception:
        return None


def extract_ipynb_text(path: Path) -> str | None:
    """Extract a notebook into a Markdown rendition: each cell as a section."""
    try:
        import json
        nb = json.loads(path.read_text(encoding="utf-8", errors="replace"))
    except Exception:
        return None
    parts: list[str] = []
    for i, cell in enumerate(nb.get("cells", []), start=1):
        ctype = cell.get("cell_type", "raw")
        src = cell.get("source", "")
        if isinstance(src, list):
            src = "".join(src)
        if not src.strip():
            continue
        if ctype == "markdown":
            parts.append(f"### Cell {i} (markdown)\n\n{src.strip()}")
        elif ctype == "code":
            lang = (nb.get("metadata", {}).get("kernelspec", {}).get("language", "python"))
            parts.append(f"### Cell {i} (code)\n\n```{lang}\n{src.rstrip()}\n```")
        else:
            parts.append(f"### Cell {i} ({ctype})\n\n{src.strip()}")
    return "\n\n".join(parts) if parts else None


def extract_csv_preview(path: Path, max_rows: int = 50) -> str | None:
    """Render a CSV/TSV as a markdown table (capped)."""
    try:
        import csv as _csv
        delim = "\t" if path.suffix.lower() == ".tsv" else ","
        with path.open("r", encoding="utf-8", errors="replace") as f:
            reader = _csv.reader(f, delimiter=delim)
            rows = []
            for i, row in enumerate(reader):
                rows.append(row)
                if i >= max_rows:
                    break
        if not rows:
            return None
        widths = [max(len(str(r[i])) for r in rows if i < len(r)) for i in range(len(rows[0]))]
        widths = [min(w, 40) for w in widths]
        lines = []
        header = rows[0]
        sep = ["-" * w for w in widths]
        lines.append("| " + " | ".join(str(h)[:40].ljust(w) for h, w in zip(header, widths)) + " |")
        lines.append("| " + " | ".join(sep) + " |")
        for r in rows[1:]:
            cells = [str(r[i])[:40].ljust(w) if i < len(r) else " " * w
                     for i, w in enumerate(widths)]
            lines.append("| " + " | ".join(cells) + " |")
        out = "\n".join(lines)
        try:
            with path.open("r", encoding="utf-8", errors="replace") as f:
                row_count = sum(1 for _ in f)
        except Exception:
            row_count = len(rows)
        if row_count > max_rows + 1:
            out += f"\n\n_Preview only — full file has approximately {row_count - 1:,} data rows._"
        return out
    except Exception:
        return None


def extract_xlsx_preview(path: Path, max_rows_per_sheet: int = 30) -> str | None:
    """Render each sheet of an xlsx as a markdown table."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    except Exception:
        return None
    parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            rows.append([("" if v is None else str(v)) for v in row])
            if i >= max_rows_per_sheet:
                break
        if not rows:
            continue
        ncols = max(len(r) for r in rows)
        rows = [r + [""] * (ncols - len(r)) for r in rows]
        widths = [min(40, max(len(rows[i][c]) for i in range(len(rows)))) for c in range(ncols)]
        header = rows[0]
        sep = ["-" * w for w in widths]
        lines = [f"### Sheet: {sheet_name}", ""]
        lines.append("| " + " | ".join(h[:40].ljust(w) for h, w in zip(header, widths)) + " |")
        lines.append("| " + " | ".join(sep) + " |")
        for r in rows[1:]:
            cells = [r[i][:40].ljust(w) for i, w in enumerate(widths)]
            lines.append("| " + " | ".join(cells) + " |")
        parts.append("\n".join(lines))
    return "\n\n".join(parts) if parts else None


def extract_text(path: Path, ext: str) -> str | None:
    if ext == "pdf":
        return extract_pdf_text(path)
    if ext in ("docx", "doc"):
        return extract_docx_text(path)
    if ext in ("txt", "md", "py", "js", "ts", "yaml", "yml"):
        try:
            content = path.read_text(encoding="utf-8", errors="replace")
            if ext in ("py", "js", "ts"):
                lang = {"py": "python", "js": "javascript", "ts": "typescript"}[ext]
                return f"```{lang}\n{content.rstrip()}\n```"
            if ext in ("yaml", "yml"):
                return f"```yaml\n{content.rstrip()}\n```"
            return content
        except Exception:
            return None
    if ext == "ipynb":
        return extract_ipynb_text(path)
    if ext in ("csv", "tsv"):
        return extract_csv_preview(path)
    if ext in ("xlsx", "xls"):
        return extract_xlsx_preview(path)
    if ext == "json":
        try:
            import json
            obj = json.loads(path.read_text(encoding="utf-8", errors="replace"))
            return f"```json\n{json.dumps(obj, indent=2, ensure_ascii=False)[:8000]}\n```"
        except Exception:
            try:
                return path.read_text(encoding="utf-8", errors="replace")
            except Exception:
                return None
    return None


# ----------------------------------------------------------------------------
# OCR fallback (Task: scanned permits / stamped drawings / signed PDFs)
# ----------------------------------------------------------------------------
def ocr_pdf(path: Path, max_pages: int = 40, dpi: int = 200):
    """OCR a scanned PDF. Returns (text|None, reason|None).

    reason is a short machine-ish string when text could not be produced, so the
    ingestion report can tell the user *why* (libs missing vs genuinely blank).
    """
    try:
        from pdf2image import convert_from_path
        import pytesseract
    except Exception:
        return None, "ocr-libs-missing (pip install pdf2image pytesseract + tesseract/poppler)"
    try:
        images = convert_from_path(str(path), dpi=dpi, first_page=1, last_page=max_pages)
    except Exception as e:
        return None, f"pdf-render-failed: {e.__class__.__name__}"
    parts = []
    for i, img in enumerate(images, start=1):
        try:
            t = pytesseract.image_to_string(img)
        except Exception as e:
            return None, f"tesseract-failed: {e.__class__.__name__}"
        if t and t.strip():
            parts.append(f"### Page {i} (OCR)\n\n{t.strip()}")
    if not parts:
        return None, "ocr-no-text (likely a non-text scan, e.g. a pure drawing)"
    return "\n\n".join(parts), None


def ocr_image(path: Path):
    """OCR a raster image. Returns (text|None, reason|None)."""
    try:
        import pytesseract
        from PIL import Image
    except Exception:
        return None, "ocr-libs-missing (pip install pytesseract pillow + tesseract)"
    try:
        with Image.open(path) as im:
            t = pytesseract.image_to_string(im)
    except Exception as e:
        return None, f"image-ocr-failed: {e.__class__.__name__}"
    if t and t.strip():
        return t.strip(), None
    return None, "ocr-no-text"


# Map the concrete extraction route to a human label for the report.
_METHOD_LABEL = {
    "native": "text layer",
    "ocr": "OCR",
    "preview": "table preview",
    "code": "code/structured",
    "copy": "verbatim copy",
}


def extract_text_rich(path: Path, ext: str, ocr: bool = True) -> dict:
    """Unified extraction with provenance. Returns a dict:
        {"text": str|None, "method": str|None, "reason": str|None}

    method ∈ {native, ocr, preview, code, copy}. When text is None, reason
    explains why so the ingestion report is actionable rather than silent.
    """
    if ext == "pdf":
        # Guard: very large PDFs (e.g. multi-MB scanned/translated books) can
        # take minutes to text-extract. Treat them as source-only so the run
        # never stalls; the _Sources stub still links the raw file.
        try:
            if path.stat().st_size > LARGE_PDF_BYTES:
                return {"text": None, "method": None,
                        "reason": f"large-pdf-skipped (>{LARGE_PDF_BYTES // (1024*1024)}MB)"}
        except Exception:
            pass
        t = extract_pdf_text(path)
        if t and t.strip():
            return {"text": t, "method": "native", "reason": None}
        if ocr:
            ot, oreason = ocr_pdf(path)
            if ot:
                return {"text": ot, "method": "ocr", "reason": None}
            return {"text": None, "method": None, "reason": oreason or "pdf-no-text"}
        return {"text": None, "method": None, "reason": "pdf-no-text-layer (OCR disabled)"}

    if ext in ("docx", "doc"):
        t = extract_docx_text(path)
        if t and t.strip():
            return {"text": t, "method": "native", "reason": None}
        return {"text": None, "method": None, "reason": "docx-empty-or-unreadable"}

    if ext in OCR_IMAGE:
        if not ocr:
            return {"text": None, "method": None, "reason": "image (OCR disabled)"}
        ot, oreason = ocr_image(path)
        if ot:
            return {"text": ot, "method": "ocr", "reason": None}
        return {"text": None, "method": None, "reason": oreason or "image-no-text"}

    # Everything else routes through the existing typed extractors.
    t = extract_text(path, ext)
    if t and t.strip():
        if ext in ("csv", "tsv", "xlsx", "xls"):
            method = "preview"
        elif ext in ("py", "js", "ts", "yaml", "yml", "json", "ipynb"):
            method = "code"
        else:
            method = "copy"
        return {"text": t, "method": method, "reason": None}
    return {"text": None, "method": None, "reason": f"{ext}-empty-or-unreadable"}


# ----------------------------------------------------------------------------
# Optional hybrid LLM concept refinement (--use-ai). Off by default so the
# default run stays fully local/offline — honoring the privacy promise that the
# client's documents never leave their machine unless they opt in.
# ----------------------------------------------------------------------------
_AI_WARNED = False  # warn only once per run if the AI layer is unavailable


def refine_concepts_ai(candidates: list[str], excerpt: str, filename: str,
                       model: str) -> tuple[list[str] | None, str | None]:
    """Ask Claude to clean the rule-based candidate list into canonical concepts.

    Returns (concepts|None, reason|None). Never raises — any failure degrades to
    the rule-based list. Requires the `anthropic` SDK and an ANTHROPIC_API_KEY.
    For the privacy promise on the landing page, point the SDK at a
    zero-data-retention endpoint/DPA when running on real client documents.
    """
    try:
        import anthropic
    except Exception:
        return None, "anthropic-sdk-missing (pip install anthropic)"
    key = os.environ.get("ANTHROPIC_API_KEY")
    if not key:
        return None, "no ANTHROPIC_API_KEY in environment"
    cand_str = ", ".join(candidates) if candidates else "(none found by rules)"
    prompt = (
        "You curate a knowledge graph. From the document excerpt below, return "
        "the 5-12 most useful CONCEPT or ENTITY names to use as note links — "
        "proper nouns, people, organizations, projects, places, regulations, "
        "and distinctive technical terms. Prefer canonical Title Case names. "
        "Drop boilerplate, dates, money amounts, and generic words. "
        "Return ONE concept per line, nothing else.\n\n"
        f"Filename: {filename}\n"
        f"Rule-based candidates: {cand_str}\n\n"
        f"Excerpt:\n{excerpt[:4000]}"
    )
    try:
        client = anthropic.Anthropic(api_key=key)
        msg = client.messages.create(
            model=model,
            max_tokens=300,
            messages=[{"role": "user", "content": prompt}],
        )
        raw = "".join(
            block.text for block in msg.content if getattr(block, "type", "") == "text"
        )
        concepts = []
        for line in raw.splitlines():
            c = line.strip().lstrip("-*•0123456789. ").strip()
            if c and len(c) <= 60:
                concepts.append(c)
        # de-dup preserving order
        seen = set()
        uniq = []
        for c in concepts:
            if c.lower() not in seen:
                seen.add(c.lower())
                uniq.append(c)
        return (uniq or None), (None if uniq else "ai-returned-empty")
    except Exception as e:
        return None, f"ai-call-failed: {e.__class__.__name__}"


# Common stopwords (EN + PT) that should never start or stand as a "concept".
# Kept deliberately small and literal — this is a rule layer, not an NLP model.
_CONCEPT_STOPWORDS = {
    # English
    "the", "a", "an", "and", "or", "but", "if", "then", "this", "that", "these",
    "those", "with", "without", "from", "into", "for", "of", "to", "in", "on",
    "at", "by", "as", "is", "are", "was", "were", "be", "been", "it", "its",
    "we", "you", "they", "he", "she", "i", "our", "your", "their", "page",
    "pages", "section", "note", "notes", "document", "figure", "table", "chapter",
    "appendix", "introduction", "conclusion", "summary", "abstract", "contents",
    # Portuguese
    "o", "os", "a", "as", "um", "uma", "uns", "umas", "de", "do", "da", "dos",
    "das", "no", "na", "nos", "nas", "em", "por", "para", "com", "sem", "que",
    "e", "ou", "se", "ao", "aos", "à", "às", "este", "esta", "esse", "essa",
    "pagina", "página", "seccao", "secção", "documento", "figura", "tabela",
    # Generic invoice/boilerplate junk seen in the wild
    "dados", "total", "bens", "atcud", "iva", "transmitente", "prestador",
    "cliente", "data", "valor", "nif", "morada", "email", "telefone",
}


def _is_good_concept(phrase: str) -> bool:
    """Filter for a phrase to qualify as a knowledge-graph concept."""
    words = phrase.split()
    if not words:
        return False
    low = phrase.lower()
    # reject if the whole phrase or its first word is a stopword
    if low in _CONCEPT_STOPWORDS or words[0].lower() in _CONCEPT_STOPWORDS:
        return False
    # reject single all-caps tokens (headers, acronyms handled separately) and
    # very short single words
    if len(words) == 1:
        w = words[0]
        if w.isupper():          # bare acronym / shout-case header
            return False
        if len(w) < 4:           # too short to be distinctive
            return False
    # reject phrases that are mostly digits/punctuation
    alpha = sum(c.isalpha() for c in phrase)
    if alpha < max(3, len(phrase) // 2):
        return False
    return True


def extract_keywords(text: str, k: int = 15) -> list[str]:
    """Extract distinctive capitalized phrases to seed a note's wikilinks.

    Rule-based candidate generator (the first half of the hybrid pipeline):
      * 1-3 capitalized words per phrase,
      * stopword + numeric + boilerplate filtering,
      * case-insensitive dedup (keeps the most common surface form),
      * ranked by frequency, then by phrase length (multi-word preferred).
    The optional --use-ai pass refines this list further; see refine_concepts_ai.
    """
    if not text:
        return []
    counts: Counter = Counter()
    forms: dict[str, Counter] = defaultdict(Counter)  # canonical->surface forms
    for line in text.splitlines():
        for m in re.finditer(
            r"\b(?:[A-ZÀ-Ú][\wÀ-ÿ\-]{2,}[ \t]+){0,2}[A-ZÀ-Ú][\wÀ-ÿ\-]{2,}\b", line
        ):
            phrase = re.sub(r"\s+", " ", m.group(0)).strip()
            if not phrase or not _is_good_concept(phrase):
                continue
            key = phrase.lower()
            counts[key] += 1
            forms[key][phrase] += 1
    if not counts:
        return []
    # Rank: frequency desc, then prefer multi-word, then alpha for stability.
    ranked = sorted(
        counts.keys(),
        key=lambda key: (-counts[key], -len(key.split()), key),
    )
    out = []
    for key in ranked[:k]:
        # restore the most common surface capitalization
        surface = forms[key].most_common(1)[0][0]
        out.append(surface)
    return out


def first_lines(text: str, n: int = 8) -> str:
    if not text: return ""
    lines = [l.strip() for l in text.splitlines() if l.strip()][:n]
    return "\n".join(lines)


def build_source_stub(src: Path, sources_dir: Path, raw_root: Path, project_tag: str, raw_root_label: str, note_link_stem: str | None = None):
    rel = src.relative_to(raw_root)
    ext = src.suffix.lower().lstrip(".")
    kind, emoji, kind_desc = kind_for(ext)
    macos_path = to_macos_path(src)
    uri = file_uri(macos_path)
    size = src.stat().st_size
    size_h = (
        f"{size/1024/1024:.1f} MB" if size >= 1024*1024
        else f"{size/1024:.0f} KB" if size >= 1024
        else f"{size} B"
    )
    category = rel.parent.parts[0] if rel.parent.parts else raw_root_label
    base = safe(src.stem)
    candidate = base
    n = 2
    while (sources_dir / f"{candidate}.md").exists() and (sources_dir / f"{candidate}.md").read_text(encoding="utf-8", errors="replace").find(macos_path) < 0:
        candidate = f"{base} ({n})"
        n += 1
    body = []
    body.append("---")
    body.append("type: source")
    body.append(f"kind: {kind}")
    body.append(f'filename: "{src.name.replace(chr(34), chr(39))}"')
    body.append(f"extension: {ext}")
    body.append(f'raw_path: "{macos_path}"')
    body.append(f"raw_path_uri: {uri}")
    body.append(f'raw_relative: "RAW /{raw_root_label}/{rel}"')
    body.append(f'category: "{category}"')
    body.append(f"size_bytes: {size}")
    body.append("tags:")
    body.append("  - source")
    body.append(f"  - source/{kind}")
    body.append(f"  - {project_tag}")
    body.append("---")
    body.append("")
    body.append(f"# {emoji} {src.name}")
    body.append("")
    body.append(f"**{kind_desc}** · {size_h} · `{category}/`")
    body.append("")
    body.append(f"[Open in macOS]({uri})")
    body.append("")
    body.append("```")
    body.append(macos_path)
    body.append("```")
    if note_link_stem:
        body.append("")
        body.append("## Extracted note")
        body.append("")
        body.append(f"- [[{note_link_stem}]]")
    (sources_dir / f"{candidate}.md").write_text("\n".join(body), encoding="utf-8")
    return candidate


def build_extracted_note(src: Path, notes_dir: Path, content: str, source_stub_stem: str, project_tag: str, raw_root_label: str, overwrite: bool = False, concepts: list[str] | None = None, method: str | None = None):
    base = safe(src.stem)
    candidate = base
    if not overwrite:
        n = 2
        while (notes_dir / f"{candidate}.md").exists():
            candidate = f"{base} ({n})"
            n += 1
    keywords = concepts if concepts else extract_keywords(content)
    excerpt = first_lines(content, 6)
    macos_path = to_macos_path(src)
    body = []
    body.append("---")
    body.append("type: extracted-note")
    body.append(f'source_filename: "{src.name.replace(chr(34), chr(39))}"')
    body.append(f'source_path: "{macos_path}"')
    body.append(f"extension: {src.suffix.lower().lstrip('.')}")
    if method:
        body.append(f"extraction_method: {method}")
    body.append("tags:")
    body.append("  - extracted")
    if method == "ocr":
        body.append("  - ocr")
    body.append(f"  - {project_tag}")
    body.append("---")
    body.append("")
    body.append(f"# {src.stem}")
    body.append("")
    body.append(f"**Source:** [[{source_stub_stem}]]")
    body.append("")
    if keywords:
        body.append("## Suggested concepts")
        body.append("")
        body.append(", ".join(f"[[{k}]]" for k in keywords))
        body.append("")
    body.append("---")
    body.append("")
    body.append("## Content")
    body.append("")
    # Cap very long extractions
    MAX_CHARS = 200_000
    if len(content) > MAX_CHARS:
        body.append(content[:MAX_CHARS])
        body.append("")
        body.append(f"_…(truncated, original is {len(content):,} chars; open the source for full content)_")
    else:
        body.append(content)
    (notes_dir / f"{candidate}.md").write_text("\n".join(body), encoding="utf-8")
    return candidate


NOTES_INDEX_START = "<!-- notes-index:start -->"
NOTES_INDEX_END = "<!-- notes-index:end -->"


def build_notes_index(vault: Path) -> list[str]:
    """Return markdown lines listing wikilinks to every note in the vault, so
    the _Index.md is a real Map of Content and the graph view connects.

    Links are vault-relative paths (e.g. [[Notes/Foo]]) to stay unambiguous
    when a stem exists in both Notes/ and _Sources/. Wrapped in marker comments
    so the block is idempotent — safe to regenerate in place.
    """
    lines = [NOTES_INDEX_START]
    for sub, heading in (("Notes", "## Notes"), ("_Sources", "## Source stubs")):
        d = vault / sub
        if not d.exists():
            continue
        stems = sorted(
            (p.relative_to(vault).with_suffix("").as_posix()
             for p in d.rglob("*.md")),
            key=str.lower,
        )
        if not stems:
            continue
        lines.append("")
        lines.append(heading)
        lines.append("")
        for rel in stems:
            label = rel.split("/")[-1]
            lines.append(f"- [[{rel}|{label}]]")
    lines.append("")
    lines.append(NOTES_INDEX_END)
    return lines


def write_index(vault: Path, vault_name: str, raw_root_label: str, sources_meta, notes_count: int, kinds: Counter, cats: Counter):
    md_count = sum(1 for _ in vault.rglob("*.md"))
    total_size = sum(s["size_bytes"] for s in sources_meta)
    lines = []
    lines.append("---")
    lines.append("type: index")
    lines.append("tags:")
    lines.append("  - moc")
    lines.append("---")
    lines.append("")
    lines.append(f"# {vault_name} — Map of Content")
    lines.append("")
    lines.append(f"**Vault notes:** {md_count:,}")
    lines.append(f"**Linked sources:** {len(sources_meta)} files (~{total_size/1024/1024:.1f} MB) under `RAW /{raw_root_label}/`")
    lines.append(f"**Extracted notes:** {notes_count}")
    lines.append("")
    lines.append("## How to use")
    lines.append("")
    lines.append("- Open the **Graph view** (Cmd+G) to see clusters.")
    lines.append("- Browse `_Sources/` to access raw files (each note has a clickable `Open in macOS` link).")
    lines.append("- Browse `Notes/` for full-text extractions of PDFs / Word docs.")
    lines.append("- The auto-generated 'Suggested concepts' wikilinks at the top of each Note are starting points; click them to spawn a new node and write your own thoughts.")
    lines.append("")
    if kinds:
        lines.append("## Sources by kind")
        lines.append("")
        for k, n in kinds.most_common():
            lines.append(f"- {k}: {n}")
        lines.append("")
    if cats:
        lines.append("## Sources by category")
        lines.append("")
        for c, n in cats.most_common():
            lines.append(f"- {c}: {n}")
        lines.append("")
    lines.extend(build_notes_index(vault))
    lines.append("")
    (vault / "_Index.md").write_text("\n".join(lines), encoding="utf-8")


def write_graph(vault: Path):
    cfg_path = vault / ".obsidian" / "graph.json"
    cfg_path.parent.mkdir(parents=True, exist_ok=True)
    cfg = {}
    if cfg_path.exists():
        try:
            cfg = json.loads(cfg_path.read_text())
        except Exception:
            cfg = {}
    cfg["showTags"] = False
    cfg["showAttachments"] = False
    cfg["hideUnresolved"] = True
    cfg.setdefault("showOrphans", False)
    cfg["collapse-color-groups"] = False
    cfg["colorGroups"] = [
        {"query": "path:_Sources/", "color": {"a": 1, "rgb": 16737095}},
        {"query": "path:Notes/", "color": {"a": 1, "rgb": 5431476}},
        {"query": "tag:#extracted", "color": {"a": 1, "rgb": 14048348}},
    ]
    cfg.setdefault("collapse-display", False)
    cfg.setdefault("textFadeMultiplier", 1.4)
    cfg.setdefault("nodeSizeMultiplier", 1.3)
    cfg.setdefault("centerStrength", 0.4)
    cfg.setdefault("repelStrength", 12)
    cfg.setdefault("linkStrength", 0.7)
    cfg.setdefault("linkDistance", 200)
    cfg.setdefault("scale", 0.15)
    cfg.setdefault("close", False)
    cfg_path.write_text(json.dumps(cfg, indent=2))


def cross_vault_hints(vault_root: Path, current: Path, source_filenames: set[str]) -> list[str]:
    """Look for matches in OTHER vaults to suggest cross-links (informational)."""
    hints = []
    for v in vault_root.iterdir():
        if not v.is_dir() or v.name.startswith("_") or v == current:
            continue
        for md in v.rglob("*.md"):
            if md.stem.lower() + md.suffix.lower() in source_filenames or md.stem.lower() in (s.rsplit(".", 1)[0] for s in source_filenames):
                hints.append(f"  {v.name}/{md.relative_to(v)}")
                if len(hints) >= 10: return hints
    return hints


def write_ingestion_report(vault: Path, vault_name: str, rows: list[dict], used_ai: bool, ai_status: str | None):
    """Write _Ingestion-Report.md — what was extracted, OCR'd, skipped, or failed.

    This is the trust layer: a non-technical client can open one note and see
    exactly what made it into the brain and what needs a second look. Returns a
    short counts dict for the console summary.
    """
    buckets = Counter(r["outcome"] for r in rows)
    failed = [r for r in rows if r["outcome"] == "failed"]
    ocrd = [r for r in rows if r["method"] == "ocr"]
    source_only = [r for r in rows if r["outcome"] == "source-only"]

    L = []
    L.append("---")
    L.append("type: ingestion-report")
    L.append("tags:")
    L.append("  - report")
    L.append("---")
    L.append("")
    L.append(f"# {vault_name} — Ingestion Report")
    L.append("")
    L.append(f"**Files seen:** {len(rows)}")
    L.append(f"**Extracted to notes:** {buckets.get('extracted', 0)}  "
             f"(of which OCR: {len(ocrd)})")
    L.append(f"**Source-only (no text):** {buckets.get('source-only', 0)}")
    L.append(f"**Failed / needs attention:** {buckets.get('failed', 0)}")
    L.append(f"**Concept engine:** {'rule-based + AI refine' if used_ai else 'rule-based'}"
             + (f" — _AI unavailable: {ai_status}; fell back to rules_" if (used_ai and ai_status) else ""))
    L.append("")

    if failed:
        L.append("## ⚠️ Needs attention — could not extract text")
        L.append("")
        L.append("| File | Category | Why |")
        L.append("| --- | --- | --- |")
        for r in sorted(failed, key=lambda r: r["filename"].lower()):
            L.append(f"| {r['filename']} | {r['category']} | {r['reason']} |")
        L.append("")

    if ocrd:
        L.append("## 🔎 Extracted via OCR (verify accuracy)")
        L.append("")
        L.append("OCR text can contain errors — spot-check these before relying on them.")
        L.append("")
        for r in sorted(ocrd, key=lambda r: r["filename"].lower()):
            L.append(f"- {r['filename']} — {r['chars']:,} chars")
        L.append("")

    if source_only:
        L.append("## 📎 Source-only (linked, not text-extracted)")
        L.append("")
        L.append("Images, CAD, BIM, media and archives are linked via `_Sources/` but carry no text layer.")
        L.append("")
        for r in sorted(source_only, key=lambda r: r["filename"].lower()):
            L.append(f"- {r['filename']} ({r['kind']})")
        L.append("")

    L.append("## Full manifest")
    L.append("")
    L.append("| File | Kind | Outcome | Method | Chars |")
    L.append("| --- | --- | --- | --- | --- |")
    for r in sorted(rows, key=lambda r: (r["outcome"], r["filename"].lower())):
        method = _METHOD_LABEL.get(r["method"], r["method"] or "—")
        chars = f"{r['chars']:,}" if r["chars"] else "—"
        L.append(f"| {r['filename']} | {r['kind']} | {r['outcome']} | {method} | {chars} |")
    L.append("")
    (vault / "_Ingestion-Report.md").write_text("\n".join(L), encoding="utf-8")
    return {
        "seen": len(rows),
        "extracted": buckets.get("extracted", 0),
        "ocr": len(ocrd),
        "source_only": buckets.get("source-only", 0),
        "failed": buckets.get("failed", 0),
    }


def resolve_routed_vault(folder_name: str, scripts_dir: Path):
    """Consult curator_routing.json to map a RAW folder name to an existing
    target vault. Returns (vault_name, rule) or (None, None) if no rule matches
    or the config is absent/malformed. Rules are tried in order; 'match' is a
    case-insensitive glob unless the rule sets 'regex': true."""
    import fnmatch
    cfg = scripts_dir / "curator_routing.json"
    if not cfg.exists():
        return None, None
    try:
        data = json.loads(cfg.read_text(encoding="utf-8"))
    except Exception as e:
        print(f"[routing] WARNING: could not parse {cfg.name}: {e}", file=sys.stderr)
        return None, None
    name = folder_name.strip()
    for rule in data.get("rules", []):
        pat = rule.get("match", "")
        if not pat:
            continue
        try:
            if rule.get("regex"):
                hit = re.search(pat, name, re.IGNORECASE) is not None
            else:
                hit = fnmatch.fnmatch(name.lower(), pat.lower())
        except re.error as e:
            print(f"[routing] WARNING: bad regex {pat!r}: {e}", file=sys.stderr)
            continue
        if hit:
            return rule.get("vault", "").strip() or None, rule
    return None, None


def main():
    ap = argparse.ArgumentParser(description="Convert RAW/<folder>/ → Vaults/<folder>/")
    ap.add_argument("folder", help="Folder name under RAW/")
    ap.add_argument("--raw", default=str(DEFAULT_RAW), help="RAW base directory")
    ap.add_argument("--vaults", default=str(DEFAULT_VAULTS), help="Vaults base directory")
    ap.add_argument("--vault-name", default=None, help="Override vault folder name (default: same as RAW folder)")
    ap.add_argument("--overwrite-notes", action="store_true", help="Re-extract Notes/ even if they already exist")
    ap.add_argument("--no-ocr", action="store_true", help="Disable OCR fallback for scanned PDFs and images")
    ap.add_argument("--use-ai", action="store_true", help="Refine concept links with an LLM pass (needs anthropic + ANTHROPIC_API_KEY; off by default to stay fully local)")
    ap.add_argument("--ai-model", default=os.environ.get("ANTHROPIC_MODEL", "claude-3-5-haiku-latest"), help="Model for --use-ai concept refinement")
    args = ap.parse_args()
    ocr_enabled = not args.no_ocr

    raw = Path(args.raw)
    vaults = Path(args.vaults)
    raw_folder = raw / args.folder
    if not raw_folder.exists():
        # If the user passed the Mac path but we're running in a Cowork sandbox,
        # the same vault lives at /sessions/<session>/mnt/Second Brain/. We
        # discover that session at runtime instead of hardcoding a session name.
        sandbox_root = None
        for candidate in Path("/sessions").glob("*/mnt/Second Brain") if Path("/sessions").exists() else []:
            if (candidate / "Vaults").exists():
                sandbox_root = candidate
                break
        if sandbox_root is not None:
            sandbox_prefix = str(sandbox_root) + "/"
            host_prefix = str(lib_vault.HOST_BASE) + "/"
            raw = Path(str(raw).replace(host_prefix, sandbox_prefix))
            vaults = Path(str(vaults).replace(host_prefix, sandbox_prefix))
            raw_folder = raw / args.folder
        if not raw_folder.exists():
            print(f"ERROR: RAW folder not found: {raw_folder}", file=sys.stderr)
            sys.exit(2)

    if args.vault_name:
        vault_name = args.vault_name.strip()
    else:
        routed, rule = resolve_routed_vault(args.folder, Path(__file__).resolve().parent)
        if routed:
            vault_name = routed
            print(f"[routing] '{args.folder}' -> existing vault '{vault_name}' "
                  f"(rule: {rule.get('match')}). Use --vault-name to override.")
        else:
            vault_name = args.folder.strip()
    vault = vaults / vault_name
    sources_dir = vault / "_Sources"
    notes_dir = vault / "Notes"
    sources_dir.mkdir(parents=True, exist_ok=True)
    notes_dir.mkdir(parents=True, exist_ok=True)

    project_tag = "project/" + re.sub(r"\s+", "-", vault_name.lower())
    raw_root_label = args.folder.strip()

    print(f"RAW:    {raw_folder}")
    print(f"VAULT:  {vault}")
    print()

    sources_meta = []
    kinds_counter = Counter()
    cats_counter = Counter()
    seen_filenames = set()
    report_rows: list[dict] = []
    n_sources = 0
    n_notes = 0
    ai_status: str | None = None  # first AI-unavailable reason, surfaced once

    files = sorted([p for p in raw_folder.rglob("*") if p.is_file() and p.name != ".DS_Store"])
    print(f"Scanning {len(files)} files…")
    if args.use_ai:
        print("AI concept refinement: ON (--use-ai)")
    for src in files:
        ext = src.suffix.lower().lstrip(".")
        kind, _, _ = kind_for(ext)
        rel_in_raw = src.relative_to(raw_folder)
        category = rel_in_raw.parent.parts[0] if rel_in_raw.parent.parts else raw_root_label
        kinds_counter[kind] += 1
        cats_counter[category] += 1
        seen_filenames.add(src.name.lower())

        # Always write the source stub (the agent contract / raw_path link).
        stub_stem = build_source_stub(src, sources_dir, raw_folder, project_tag, raw_root_label, note_link_stem=None)
        n_sources += 1

        row = {"filename": src.name, "category": category, "kind": kind,
               "ext": ext, "outcome": "source-only", "method": None,
               "reason": None, "chars": 0}

        attempt = ext in TEXT_EXTRACT or (ocr_enabled and ext in OCR_IMAGE)
        if attempt:
            res = extract_text_rich(src, ext, ocr=ocr_enabled)
            text, method, reason = res["text"], res["method"], res["reason"]
            if text and text.strip():
                target = notes_dir / f"{safe(src.stem)}.md"
                if target.exists() and not args.overwrite_notes:
                    # Note already present from a prior run; count as extracted.
                    row.update(outcome="extracted", method=method, chars=len(text))
                else:
                    # Concept links: rule-based candidates, optionally AI-refined.
                    concepts = extract_keywords(text)
                    if args.use_ai and len(text) > 200:
                        refined, why = refine_concepts_ai(
                            concepts, first_lines(text, 40), src.name, args.ai_model)
                        if refined:
                            concepts = refined
                        elif why and ai_status is None:
                            ai_status = why  # report once
                    note_stem = build_extracted_note(
                        src, notes_dir, text, stub_stem, project_tag, raw_root_label,
                        overwrite=args.overwrite_notes, concepts=concepts, method=method)
                    n_notes += 1
                    row.update(outcome="extracted", method=method, chars=len(text))
                    # Link the stub to its extracted note.
                    stub_path = sources_dir / f"{stub_stem}.md"
                    stub_text = stub_path.read_text(encoding="utf-8")
                    if "## Extracted note" not in stub_text:
                        stub_text += f"\n\n## Extracted note\n\n- [[{note_stem}]]\n"
                        stub_path.write_text(stub_text, encoding="utf-8")
            else:
                # No text. For images that's normal (source-only); for text docs
                # it's a genuine failure worth flagging.
                if ext in OCR_IMAGE:
                    row.update(outcome="source-only", method=None, reason=reason)
                else:
                    row.update(outcome="failed", method=None, reason=reason)

        report_rows.append(row)
        sources_meta.append({
            "stem": stub_stem,
            "filename": src.name,
            "kind": kind,
            "category": category,
            "size_bytes": src.stat().st_size,
        })

    write_index(vault, vault_name, raw_root_label, sources_meta, n_notes, kinds_counter, cats_counter)
    write_graph(vault)
    summary = write_ingestion_report(vault, vault_name, report_rows, args.use_ai, ai_status)

    print(f"\n✅ Done.")
    print(f"   Source stubs:    {n_sources} (in _Sources/)")
    print(f"   Extracted notes: {n_notes} (in Notes/)  [OCR: {summary['ocr']}]")
    print(f"   Source-only:     {summary['source_only']}")
    print(f"   ⚠️  Failed:       {summary['failed']}  (see _Ingestion-Report.md)")
    print(f"   Index:  {vault / '_Index.md'}")
    print(f"   Report: {vault / '_Ingestion-Report.md'}")
    print(f"   Graph:  {vault / '.obsidian/graph.json'}")
    if args.use_ai and ai_status:
        print(f"   Note: AI refine unavailable ({ai_status}); used rule-based concepts.")

    # Cross-vault hint
    hints = cross_vault_hints(vaults, vault, seen_filenames)
    if hints:
        print(f"\n🔗 Possible matches in other vaults (informational, not auto-linked):")
        for h in hints:
            print(h)


if __name__ == "__main__":
    main()
